"""
excel_generator.py
Builds the complete Sprint Report Excel file:
  - Sprint Summary block (rows 1-14) at the top
  - Full Epic → Story/Task → Subtask hierarchy table below
"""

import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

JIRA_BASE = "https://jira-zigram.atlassian.net/browse"

# ── Colour constants ──────────────────────────────────────────────────────────
WHITE   = 'FFFFFF'
BLACK   = '000000'

# Summary block header colours
C_SPRINT_NO   = '000000'
C_START       = '1F4E79'
C_DEV         = '2E75B6'
C_QA          = '00B0F0'
C_PROD        = 'C00000'
C_END         = '375623'
C_DAYS        = '595959'
C_SCRUM       = '7030A0'
C_KPI_AI      = '1F3864'
C_KPI_PEND    = 'FFC000'
C_KPI_NOTINIT = 'ED7D31'
C_KPI_PRODPCT = '375623'
C_GOAL        = '7030A0'
C_MAJOR       = '1F3864'

# Status breakdown colours
C_PENDING_AI  = 'F4B942'
C_NOTINIT     = 'ED7D31'
C_INPROG      = '00B0F0'
C_STAGING     = 'BF8F00'
C_QAREVIEW    = 'FFC000'
C_QADEP       = '70AD47'
C_QAAPP       = '00B050'
C_PRODUCTION  = '375623'
C_ONHOLD      = 'A6A6A6'
C_ANOTHER     = '7030A0'

# Hierarchy row colours
EPIC_BG   = '1F3864'
STORY_BG  = '2E75B6'
SUB_BG    = 'D9E2F3'

# Status cell colours
STATUS_DONE    = ('C6EFCE', '375623')
STATUS_INPROG  = ('FFEB9C', '7F6000')
STATUS_TODO    = ('FCE4D6', '9C0006')
STATUS_STAGING = ('FFF2CC', '7F6000')


# ── Style helpers ─────────────────────────────────────────────────────────────

def _hdr(bg, fg=WHITE, bold=True, sz=9):
    return dict(
        font      = Font(bold=bold, color=fg, name='Arial', size=sz),
        fill      = PatternFill('solid', start_color=bg),
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True),
        border    = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF'),
        )
    )


def _val(bg=WHITE, fg=BLACK, bold=False, sz=10, h='center'):
    thin = Side(style='thin', color='CCCCCC')
    return dict(
        font      = Font(bold=bold, color=fg, name='Arial', size=sz),
        fill      = PatternFill('solid', start_color=bg),
        alignment = Alignment(horizontal=h, vertical='center', wrap_text=True),
        border    = Border(left=thin, right=thin, top=thin, bottom=thin),
    )


def _apply(cell, style: dict):
    for k, v in style.items():
        setattr(cell, k, v)


def _blank_row(ws, row, cols, bg='F2F2F2', height=6):
    ws.row_dimensions[row].height = height
    for col in cols:
        ws[f'{col}{row}'].fill = PatternFill('solid', start_color=bg)


# ── Main builder ──────────────────────────────────────────────────────────────

def build_excel(form_data: dict, parsed: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Sprint {form_data['sprint_number']} - Report"

    kpis      = parsed['kpis']
    hierarchy = parsed['hierarchy']
    fd        = form_data

    daily_task = round(kpis['action_items'] / fd['total_days'], 2) \
        if fd['total_days'] > 0 else 0

    ALL_COLS   = ['A','B','C','D','E','F','G','H','I','J','K']
    META_COLS  = ['A','B','C','D','E','F','G','H']
    STAT_COLS  = ALL_COLS

    meta_hdrs = [
        ('A1', 'Sprint Number',              C_SPRINT_NO),
        ('B1', 'Sprint Start Date',          C_START),
        ('C1', 'Sprint Development Release', C_DEV),
        ('D1', 'Sprint QA Release',          C_QA),
        ('E1', 'Production Release',         C_PROD),
        ('F1', 'Sprint End Date',            C_END),
        ('G1', 'Total No. of Days',          C_DAYS),
        ('H1', 'Scrum Master',               C_SCRUM),
    ]
    for ref, label, bg in meta_hdrs:
        c = ws[ref]; c.value = label; _apply(c, _hdr(bg))
    ws.row_dimensions[1].height = 28

    meta_vals = {
        'A2': fd['sprint_number'],
        'B2': fd['sprint_start'].strftime('%d %b %Y'),
        'C2': fd['dev_release'].strftime('%d %b %Y'),
        'D2': fd['qa_release'].strftime('%d %b %Y'),
        'E2': fd['prod_release'].strftime('%d %b %Y'),
        'F2': fd['sprint_end'].strftime('%d %b %Y'),
        'G2': fd['total_days'],
        'H2': fd['scrum_master'],
    }
    for ref, val in meta_vals.items():
        c = ws[ref]; c.value = val; _apply(c, _val('FFF2CC', BLACK, True, 10))
    ws.row_dimensions[2].height = 22

    _blank_row(ws, 3, META_COLS)

    kpi_hdrs = [
        ('A4', 'No of Days Left in Sprint', C_SPRINT_NO),
        ('B4', 'Action Items',              C_KPI_AI),
        ('C4', 'Pending %',                 C_KPI_PEND),
        ('D4', 'Not Initiated %',           C_KPI_NOTINIT),
        ('E4', 'Production Release %',      C_KPI_PRODPCT),
        ('F4', '',                          'D9D9D9'),
        ('G4', '',                          'D9D9D9'),
        ('H4', '',                          'D9D9D9'),
    ]
    for ref, label, bg in kpi_hdrs:
        c = ws[ref]; c.value = label; _apply(c, _hdr(bg))
    ws.row_dimensions[4].height = 28

    kpi_vals = {
        'A5': fd['days_left'],
        'B5': kpis['action_items'],
        'C5': kpis['pending_pct'],
        'D5': kpis['not_initiated_pct'],
        'E5': kpis['production_release_pct'],
    }
    for ref, val in kpi_vals.items():
        c = ws[ref]; c.value = val; _apply(c, _val(WHITE, BLACK, True, 11))
    for col in ['F','G','H']:
        ws[f'{col}5'].fill = PatternFill('solid', start_color='F2F2F2')
    ws.row_dimensions[5].height = 22

    _blank_row(ws, 6, META_COLS)

    stat_hdrs = [
        ('A7', 'Daily Task Count',              C_DAYS),
        ('B7', 'Pending Action Items',          C_PENDING_AI),
        ('C7', 'Not Initiated',                 C_NOTINIT),
        ('D7', 'In Progress',                   C_INPROG),
        ('E7', 'Staging',                       C_STAGING),
        ('F7', 'QA Review',                     C_QAREVIEW),
        ('G7', 'QA Deployed',                   C_QADEP),
        ('H7', 'QA Approved',                   C_QAAPP),
        ('I7', 'Production',                    C_PRODUCTION),
        ('J7', 'On Hold',                       C_ONHOLD),
        ('K7', 'To Be Picked In Another Sprint',C_ANOTHER),
    ]
    for ref, label, bg in stat_hdrs:
        c = ws[ref]; c.value = label; _apply(c, _hdr(bg))
    ws.row_dimensions[7].height = 36

    stat_vals = {
        'A8': daily_task,
        'B8': kpis['pending_action_items'],
        'C8': kpis['not_initiated'],
        'D8': kpis['in_progress'],
        'E8': kpis['staging'],
        'F8': kpis['qa_review'],
        'G8': kpis['qa_deployed'],
        'H8': kpis['qa_approved'],
        'I8': kpis['production'],
        'J8': kpis['on_hold'],
        'K8': kpis['to_be_picked'],
    }
    for ref, val in stat_vals.items():
        c = ws[ref]; c.value = val; _apply(c, _val(WHITE, BLACK, True, 11))
    ws.row_dimensions[8].height = 22

    _blank_row(ws, 9, STAT_COLS)

    c = ws['A10']; c.value = 'Sprint Goal';      _apply(c, _hdr(C_GOAL))
    c = ws['B10']; c.value = 'Major Sprint Items'; _apply(c, _hdr(C_MAJOR))
    for col in ['C','D','E','F','G','H','I','J','K']:
        ws[f'{col}10'].fill = PatternFill('solid', start_color='F2F2F2')
    ws.row_dimensions[10].height = 24

    for i, (row, val_a, val_b) in enumerate([
        (11, fd['sprint_goal'],  fd['major_item_1']),
        (12, '',                 fd['major_item_2']),
        (13, '',                 fd['major_item_3']),
    ]):
        ca = ws[f'A{row}']; ca.value = val_a
        _apply(ca, _val('FAE5D3' if i == 0 else 'F2F2F2', BLACK, False, 9, 'left'))

        cb = ws[f'B{row}']; cb.value = val_b
        _apply(cb, _val('FFF2CC', BLACK, False, 9, 'left'))

        for col in ['C','D','E','F','G','H','I','J','K']:
            ws[f'{col}{row}'].fill = PatternFill('solid', start_color='F2F2F2')
        ws.row_dimensions[row].height = 20

    ws.row_dimensions[14].height = 14
    ws['A14'].value = '🟡 Yellow = Manual Input   |   Auto-calculated fields derived from Jira CSV'
    ws['A14'].font  = Font(name='Arial', size=8, italic=True, color='595959')

    task_headers = ['Issue Key', 'URL', 'Issue Type', 'Summary / Title',
                    'Status', 'Priority', 'Assignee',
                    'Projected Start', 'Projected End']
    for col_idx, label in enumerate(task_headers, 1):
        c = ws.cell(15, col_idx)
        c.value = label
        _apply(c, _hdr('1F3864'))
    ws.row_dimensions[15].height = 28

    thin_w = Side(style='thin', color='B8CCE4')
    thin_g = Side(style='thin', color='CCCCCC')
    border_epic = Border(
        left=Side(style='medium', color='1F3864'),
        right=Side(style='medium', color='1F3864'),
        top=Side(style='medium', color='1F3864'),
        bottom=Side(style='medium', color='1F3864'),
    )
    border_sub = Border(left=thin_g, right=thin_g, top=thin_g, bottom=thin_g)

    rows_with_spacers = []
    for i, item in enumerate(hierarchy):
        rows_with_spacers.append(item)
        if i + 1 < len(hierarchy) and hierarchy[i+1]['level'] == 0:
            rows_with_spacers.append(None)

    current_row = 16
    for item in rows_with_spacers:

        if item is None:
            ws.row_dimensions[current_row].height = 8
            for col in range(1, 10):
                ws.cell(current_row, col).fill = PatternFill('solid', start_color='FFFFFF')
            current_row += 1
            continue

        level   = item['level']
        ik      = item['issue_key']
        url     = f"{JIRA_BASE}/{ik}"

        if level == 0:
            summary_disp = '⬛  ' + item['summary'].upper()
            bg, fg, bold, sz = EPIC_BG, WHITE, True, 11
            row_h = 28
        elif level == 1:
            summary_disp = '    ▶  ' + item['summary']
            bg, fg, bold, sz = STORY_BG, WHITE, True, 10
            row_h = 22
        else:
            summary_disp = '         ◦  ' + item['summary']
            bg, fg, bold, sz = SUB_BG, '1F3864', False, 9
            row_h = 18

        ws.row_dimensions[current_row].height = row_h

        row_data = [
            ik, url, item['issue_type'], summary_disp,
            item['status'], item['priority'], item['assignee'],
            item['target_start'], item['target_end'],
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(current_row, col_idx)
            cell.alignment = Alignment(vertical='center', wrap_text=(col_idx == 4))
            cell.border    = border_epic if level == 0 else border_sub

            if col_idx == 2:
                cell.value     = val
                cell.hyperlink = val
                cell.font      = Font(bold=False, color='4472C4', name='Arial',
                                      size=sz, underline='single')
                cell.fill      = PatternFill('solid', start_color=bg)

            elif col_idx == 5:
                sv = str(item['status'])
                if any(s in sv for s in ['Done', 'DONE', 'Production', 'Released']):
                    sc = STATUS_DONE
                elif any(s in sv for s in ['Progress', 'IN PROGRESS']):
                    sc = STATUS_INPROG
                elif any(s in sv for s in ['Staging', 'STAGING']):
                    sc = STATUS_STAGING
                else:
                    sc = STATUS_TODO
                cell.value = val
                cell.fill  = PatternFill('solid', start_color=sc[0])
                cell.font  = Font(bold=True, name='Arial', size=sz, color=sc[1])

            else:
                cell.value = val
                cell.font  = Font(bold=bold, color=fg, name='Arial', size=sz)
                cell.fill  = PatternFill('solid', start_color=bg)

        current_row += 1

    col_widths = {
        'A': 14,
        'B': 50,
        'C': 14,
        'D': 58,
        'E': 16,
        'F': 13,
        'G': 24,
        'H': 16,
        'I': 16,
        'J': 14,
        'K': 24,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A16'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
